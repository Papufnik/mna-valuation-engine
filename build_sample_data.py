"""
build_sample_data.py -- generates the illustrative sample_data/ tax package
this repo's mna_valuation_engine.py reads a live Net Income figure from
(cell B59 of the "2 - 2025 Income Statement" tab), so a reviewer can clone
this repo and run the pipeline end to end with no real business data.
"""
import os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SAMPLE_DIR = os.path.join(HERE, "sample_data")
os.makedirs(SAMPLE_DIR, exist_ok=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "2 - 2025 Income Statement"

ws["A1"] = "Illustrative Income Statement (sample data -- not a real business)"
ws["A2"] = "Revenue"
ws["B2"] = 1142300.00
ws["A59"] = "Net Income"
ws["B59"] = 148220.60

# a couple of decoy sheets, matching the shape of the real multi-tab tax
# package this script reads a single verified cell out of.
wb.create_sheet("0 - Methodology Notes")
wb.create_sheet("3 - Balance Sheet Summary")
wb.create_sheet("Tab 10 - Related Entity Notes")

out_path = os.path.join(SAMPLE_DIR, "business_tax_package_latest.xlsx")
wb.save(out_path)
print(f"wrote {out_path}")

